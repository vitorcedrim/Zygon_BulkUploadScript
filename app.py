from email.mime import image
import string
import zipfile, glob, os, openpyxl, re
from PySimpleGUI.PySimpleGUI import WIN_CLOSED, popup
from re import search
from openpyxl import Workbook
import PySimpleGUI as sg
from PIL import Image
from os import listdir
from os.path import isfile, join
import pathlib
import csv

varTaxonomy = ""
varCreativeType = ""
varClickTroughURL = ""
varTrackingURL = ""

currentDirectory = str(os.path.dirname(os.path.realpath(__file__)))
print(currentDirectory)

imageFiles = [] #Armazena os nomes dos arquivos de imagem.
imageSizes = []

os.chdir(currentDirectory)
for file in glob.glob("*.png"): #Coleta os nomes dos arquivos de imagem PNG.
    imageFiles.append(file)

for file in glob.glob("*.jpg"): #Coleta os nomes dos arquivos de imagem JPG.
    imageFiles.append(file)
    
for file in glob.glob("*.jpeg"): #Coleta os nomes dos arquivos de imagem JPEG.
    imageFiles.append(file)
    
for file in glob.glob("*.gif"): #Coleta os nomes dos arquivos de imagem GIF.
    imageFiles.append(file)
    
print(imageFiles)



for imageIndex in imageFiles:
    image = Image.open(f'{currentDirectory}\{imageIndex}')
    width = image.width
    height = image.height
    imageSizes.append(f'{width}x{height}')
    
print (imageSizes)

def displayCreativeWorkflow():
    wb = Workbook()
    ws = wb.active

    arrayFileList = imageFiles #Armazena os nomes dos arquivos de imagem.
    arrayImageSizes = imageSizes #Armazena as dimensões dos arquivos de imagem.
    

    varCreativeQuantity = 0 #Quantidade de criativos dentro do ZIP, utilizado como condição para interromper o loop.
    varCreativeNumber = 0 #Número/índice do criativo sendo tratado no loop atual.
    varFileName = "" #Nome do arquivo de criativo sendo tratado no loop atual.
    varFileSize = "" #Tamanho em pixels do criativo sendo tratado no loop atual.
    varCreativeName = "" #Nome do arquivo de criativo sendo tratado no loop atual.



    varCreativeQuantity = len(arrayFileList)

    def createDisplayTemplate():
        ws.title = "DV360 - Display" #Define o título do documento CSV.
        
        ws['A1'] = "Creative name"
        ws['B1'] = "Main asset file name"
        ws['C1'] = "Backup image file name (for HTML5 only)"
        ws['D1'] = "Click-through URL"
        ws['E1'] = "Dimensions (width x height)"
        ws['F1'] = "Appended HTML tag (Optional)"
        ws['G1'] = "Integration code (Optional)"
        ws['H1'] = "Notes (Optional)"

    def writeDataDisplay():
        ws[f'A{varCreativeNumber + 2}'] = varCreativeName
        ws[f'B{varCreativeNumber + 2}'] = varFileName
        ws[f'D{varCreativeNumber + 2}'] = varClickTroughURL
        ws[f'E{varCreativeNumber + 2}'] = arrayImageSizes[varCreativeNumber]
        ws[f'F{varCreativeNumber + 2}'] = varTrackingURL

    createDisplayTemplate()


    while varCreativeNumber < varCreativeQuantity:
        varFileName = arrayFileList[varCreativeNumber] #Coloca o nome do arquivo baseado no arquivo que está sendo tratado no loop atual.
        varCreativeName = varTaxonomy + arrayImageSizes[varCreativeNumber]
        print(varCreativeName)
        writeDataDisplay() #Escreve os dados na planilha.
        varCreativeNumber += 1 #Avança o índice de criativo a ser tratado.
        
    with open('DV360 - Display.csv', 'w', newline="") as file_handle:
        csv_writer = csv.writer(file_handle)
        for row in ws.iter_rows(): # generator; was sh.rows
            csv_writer.writerow([cell.value for cell in row])

def taxonomyWindow():
    sg.theme("Reddit")
    layout = [
        [sg.Text("Insira a taxonomia Zygon para criativo (Sem dimensões).")],
        [sg.Input(key="taxonomy")],
        [sg.Button("Continuar")]
    ]
    return sg.Window("Taxonomia", layout=layout,finalize=True)

def extraInfosWindow():
    sg.theme("Reddit")
    layout = [
        [sg.Text("Insira as informações extras.")],
        [sg.Text("Click Through URL (Lembre de parametrizar)")],[sg.Input(key="clickTroughURL")],
        [sg.Text("Tracking URL (Opcional)")],[sg.Input(key="trackingURL")],
        [sg.Button("Criar planilha")]
    ]
    return sg.Window("Extras", layout=layout,finalize=True)

def finalWindow():
    sg.theme("Reddit")
    layout = [
        [sg.Text("Script criado com muito carinho para a equipe de operações Zygon por:")],
        [sg.Text("Vitor Cedrim")],
        [sg.Button("Fechar")]
    ]
    return sg.Window("Obrigado!", layout=layout,finalize=True, element_justification='center')

varTaxonomyWindow = taxonomyWindow()
varCreativeTypeWindow = None
varExtraInfosWindow = None
varFinalWindow = None

while True:
    window,event,values = sg.read_all_windows()

    if window == varTaxonomyWindow and event == sg.WIN_CLOSED:
        break

    if window == varTaxonomyWindow and event == "Continuar" and values["taxonomy"] != "":
        varTaxonomy = values["taxonomy"]
        varExtraInfosWindow = extraInfosWindow()
        varTaxonomyWindow.hide()
    if window == varExtraInfosWindow and event == sg.WIN_CLOSED:
        break
    if window == varExtraInfosWindow and event == "Criar planilha":
        varClickTroughURL = values["clickTroughURL"]
        varTrackingURL = values["trackingURL"]
        displayCreativeWorkflow()
        varExtraInfosWindow.hide()
        varFinalWindow = finalWindow()
    if window == varFinalWindow and event == "Fechar":
        break
    elif window == varFinalWindow and event == sg.WIN_CLOSED:
        break